import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.model_selection import train_test_split
from sklearn.svm import SVC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Excel dosyasını oku ve bir DataFrame'e yükle
veri_seti = pd.read_excel('/Users/hakanmartin/PycharmProjects/pythonProject1/ProjeVeriSetiEn_TR.xlsx')

# Türkçe metinlerin olduğu 'TR' sütununu seçin
metinler = veri_seti['TR'].values.astype('U')

# TF-IDF vektörleştiriciyi tanımlayın ve uygulayın
tfidf_vectorizer = TfidfVectorizer()
tfidf_vectors = tfidf_vectorizer.fit_transform(metinler)

# TF-IDF vektörlerini X, niyet etiketlerini y olarak tanımlama
X = tfidf_vectors
y = veri_seti['INTENT']

# Eğitim ve test veri setlerine ayırma
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Support Vector Machine modeli oluşturma ve eğitme
svm_model = SVC(kernel='linear')
svm_model.fit(X_train, y_train)

# Test veri seti üzerinde tahmin yapma
y_pred = svm_model.predict(X_test)

# Tahmin edilen niyetler ve gerçek etiketlerin olduğu bir DataFrame oluşturma
tahmin_df = pd.DataFrame({'Gerçek Etiket': y_test, 'Tahmin Edilen Niyet': y_pred})

# Excel dosyasına yazma
tahmin_df.to_excel('/Users/hakanmartin/PycharmProjects/pythonProject1/bos.xlsx', index=False)

# Yanlış tahmin edilen niyetleri işaretlemek için Excel dosyasını açma
workbook = load_workbook(filename='/Users/hakanmartin/PycharmProjects/pythonProject1/bos.xlsx')
sheet = workbook.active

# Yanlış tahmin edilen niyetleri işaretlemek için hücreleri kontrol etme
for row in sheet.iter_rows(min_row=2, max_row=len(y_pred) + 1, min_col=1, max_col=2):
    gercek_niyet = row[0].value
    tahmin_edilen_niyet = row[1].value

    if tahmin_edilen_niyet != gercek_niyet:
        row[1].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Excel dosyasını kaydetme
workbook.save(filename='/Users/hakanmartin/PycharmProjects/pythonProject1/bos.xlsx')
